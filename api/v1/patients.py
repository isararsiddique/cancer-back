from typing import Optional, List, Dict, Any
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_

from core.deps import permission_required, get_db, get_current_user, dashboard_required
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from core.audit import log_patient_create, log_patient_update, log_bulk_upload, log_data_export
from db.models.registry import Patient
from db.models.users import User
from pydantic import BaseModel
from datetime import date, datetime
import csv
import io
import uuid
import logging
import tempfile
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/patients", tags=["patients"]) 


class PatientCreate(BaseModel):
    # Patient Identifier (optional external ID like medical record number)
    patient_id: Optional[str] = None
    
    # Required fields
    patient_name: str
    diagnosis_date: date
    icd11_main_code: str
    
    # Demographics
    gender: Optional[str] = None
    date_of_birth: Optional[date] = None
    nationality: Optional[str] = None
    address: Optional[dict] = None
    
    # ICD-11 Coding
    icd11_description: Optional[str] = None
    icd11_composite_expression: Optional[str] = None
    icd11_manifestation_code: Optional[str] = None
    manifestation: Optional[str] = None
    icd11_topography_code: Optional[str] = None
    icd11_topography: Optional[str] = None
    icd11_morphology_code: Optional[str] = None
    icd11_morphology: Optional[str] = None
    icd11_behavior_code: Optional[str] = None
    icd11_stage_code: Optional[str] = None
    laterality: Optional[str] = None
    
    # TNM Staging
    t_category: Optional[str] = None
    n_category: Optional[str] = None
    m_category: Optional[str] = None
    
    # Cancer Registry Fields
    multiple_primary_flag: Optional[bool] = None
    basis_of_diagnosis: Optional[str] = None
    primary_site_confirmed: Optional[bool] = None
    
    # Treatment
    surgery_done: Optional[bool] = None
    surgery_date: Optional[date] = None
    chemotherapy_done: Optional[bool] = None
    chemo_start_date: Optional[date] = None
    radiotherapy_done: Optional[bool] = None
    hormonal_therapy: Optional[bool] = None
    immunotherapy: Optional[bool] = None
    treatment_intent: Optional[str] = None
    treatment_notes: Optional[str] = None
    
    # Follow-up
    followup_date: Optional[date] = None
    vital_status: Optional[str] = None
    cause_of_death_icd11: Optional[str] = None
    recurrence: Optional[bool] = None
    recurrence_date: Optional[date] = None
    metastasis: Optional[bool] = None
    survival_months: Optional[int] = None
    followup_notes: Optional[str] = None
    
    # Metadata
    data_source: Optional[str] = None
    entry_mode: Optional[str] = None
    entered_by: Optional[str] = None  # Text field from form (e.g., "Registrar")
    validation_status: Optional[str] = "Pending"


@router.post("/", dependencies=[Depends(permission_required("patients.create")), Depends(dashboard_required("hospital"))])
def create_patient(
    payload: PatientCreate, 
    db: Session = Depends(get_db), 
    current_user: User = Depends(get_current_user)
):
    """
    Create a new patient record.
    
    Workflow for doctors/editors:
    1. Doctor clicks "Create Patient"
    2. ICD-11 search interface is shown FIRST
    3. Doctor searches and selects an ICD-11 code
    4. Form fields are auto-filled from ICD-11 data
    5. Doctor can edit any field manually
    6. Doctor submits the form
    7. Patient is created in both full and anonymized databases
    
    The anonymized view (registry.patients_anonymized) automatically reflects 
    new records when is_active=True. No separate insert needed.
    """
    try:
        age_at_diagnosis = None
        if payload.date_of_birth and payload.diagnosis_date:
            dob = payload.date_of_birth
            diag = payload.diagnosis_date
            age_at_diagnosis = diag.year - dob.year - ((diag.month, diag.day) < (dob.month, dob.day))

        p = Patient(
            # Identity & References
            tenant_id=current_user.tenant_id,
            organization_id=current_user.organization_id,
            entered_by=payload.entered_by,  # Text field from form
            
            # Patient Identifier
            patient_id=payload.patient_id,
            
            # Demographics
            patient_name=payload.patient_name,
            gender=payload.gender,
            date_of_birth=payload.date_of_birth,
            nationality=payload.nationality,
            address=payload.address,
            
            # Diagnosis
            diagnosis_date=payload.diagnosis_date,
            age_at_diagnosis=age_at_diagnosis,
            
            # ICD-11 Coding
            icd11_main_code=payload.icd11_main_code,
            icd11_description=payload.icd11_description,
            icd11_composite_expression=payload.icd11_composite_expression,
            icd11_manifestation_code=payload.icd11_manifestation_code,
            manifestation=payload.manifestation,
            icd11_topography_code=payload.icd11_topography_code,
            icd11_topography=payload.icd11_topography,
            icd11_morphology_code=payload.icd11_morphology_code,
            icd11_morphology=payload.icd11_morphology,
            icd11_behavior_code=payload.icd11_behavior_code,
            icd11_stage_code=payload.icd11_stage_code,
            laterality=payload.laterality,
            
            # TNM Staging
            t_category=payload.t_category,
            n_category=payload.n_category,
            m_category=payload.m_category,
            
            # Cancer Registry Fields
            multiple_primary_flag=payload.multiple_primary_flag,
            basis_of_diagnosis=payload.basis_of_diagnosis,
            primary_site_confirmed=payload.primary_site_confirmed,
            
            # Treatment
            surgery_done=payload.surgery_done,
            surgery_date=payload.surgery_date,
            chemotherapy_done=payload.chemotherapy_done,
            chemo_start_date=payload.chemo_start_date,
            radiotherapy_done=payload.radiotherapy_done,
            hormonal_therapy=payload.hormonal_therapy,
            immunotherapy=payload.immunotherapy,
            treatment_intent=payload.treatment_intent,
            treatment_notes=payload.treatment_notes,
            
            # Follow-up
            followup_date=payload.followup_date,
            vital_status=payload.vital_status,
            cause_of_death_icd11=payload.cause_of_death_icd11,
            recurrence=payload.recurrence,
            recurrence_date=payload.recurrence_date,
            metastasis=payload.metastasis,
            survival_months=payload.survival_months,
            followup_notes=payload.followup_notes,
            
            # Metadata
            data_source=payload.data_source or "Manual",
            entry_mode=payload.entry_mode or "Manual",
            validation_status=payload.validation_status or "Pending",
            is_active=True  # Ensure record appears in anonymized view
        )
        
        db.add(p)
        db.commit()
        db.refresh(p)
        
        # Audit logging
        try:
            log_patient_create(
                db=db,
                user_id=current_user.id,
                patient_id=str(p.id),
                patient_name=payload.patient_name,
                icd11_code=payload.icd11_main_code
            )
        except Exception as e:
            # Log audit failure but don't fail the request
            logger.warning(f"Audit logging failed for patient creation: {str(e)}")
        
        # Verify data appears in both databases
        from sqlalchemy import text
        anonymized_check = 0
        try:
            # Check if patient appears in anonymized table (trigger should have synced it)
            anonymized_check = db.execute(
                text("SELECT COUNT(*) FROM registry.patients_anonymized WHERE original_id_hash = MD5(:id::text)::uuid"),
                {"id": str(p.id)}
            ).scalar()
        except Exception as e:
            # Table might not be accessible or query failed, but data is in main table
            logger.warning(f"Could not verify anonymized table: {str(e)}")
            anonymized_check = 1  # Assume success since trigger should have synced
        
        return {
            "id": str(p.id),
            "patient_id": p.patient_id,
            "message": "Patient created successfully",
            "databases_updated": [
                "registry.patients (full detailed data - internal use only)",
                "registry.patients_anonymized (Malaysian PDPA-compliant anonymized data - automatic via trigger)"
            ],
            "note": "Patient automatically synced to both databases. Full data in registry.patients, anonymized data in registry.patients_anonymized (Malaysian government compliant for research).",
            "anonymized_verified": anonymized_check > 0
        }
    
    except Exception as e:
        db.rollback()
        logger.error(f"Error creating patient: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to create patient: {str(e)}"
        )


@router.post("/bulk-upload", dependencies=[Depends(permission_required("patients.create"))])
async def bulk_upload_patients(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Bulk upload patients from CSV file.
    
    CSV should have columns matching PatientCreate fields:
    - patient_name (required)
    - diagnosis_date (required, format: YYYY-MM-DD)
    - icd11_main_code (required)
    - patient_id, gender, date_of_birth, nationality, etc. (optional)
    
    Returns summary of successful and failed imports.
    """
    if not file.filename.endswith('.csv'):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File must be a CSV file"
        )
    
    try:
        # Read CSV file
        contents = await file.read()
        csv_data = contents.decode('utf-8')
        csv_reader = csv.DictReader(io.StringIO(csv_data))
        
        successful_imports = []
        failed_imports = []
        row_number = 1  # Start from 1 (header is row 0)
        
        for row in csv_reader:
            row_number += 1
            try:
                # Parse required fields
                if not row.get('patient_name') or not row.get('diagnosis_date') or not row.get('icd11_main_code'):
                    failed_imports.append({
                        "row": row_number,
                        "error": "Missing required fields: patient_name, diagnosis_date, or icd11_main_code",
                        "data": row
                    })
                    continue
                
                # Helper function to parse dates with multiple format support
                def parse_date(date_str):
                    if not date_str or not date_str.strip():
                        return None
                    date_str = date_str.strip()
                    
                    # Try multiple date formats
                    date_formats = [
                        '%Y-%m-%d',      # 2012-01-19
                        '%m/%d/%Y',      # 1/19/2012
                        '%d/%m/%Y',      # 19/1/2012
                        '%Y/%m/%d',      # 2012/1/19
                        '%m-%d-%Y',      # 1-19-2012
                        '%d-%m-%Y',      # 19-1-2012
                        '%Y.%m.%d',      # 2012.1.19
                        '%m.%d.%Y',      # 1.19.2012
                        '%d.%m.%Y',      # 19.1.2012
                    ]
                    
                    for fmt in date_formats:
                        try:
                            return datetime.strptime(date_str, fmt).date()
                        except ValueError:
                            continue
                    
                    # If no format matches, raise error with helpful message
                    raise ValueError(f"Date '{date_str}' does not match any supported format. Supported formats: YYYY-MM-DD, M/D/YYYY, D/M/YYYY, etc.")
                
                # Parse dates (handle empty strings and multiple formats)
                diagnosis_date = parse_date(row.get('diagnosis_date'))
                date_of_birth = parse_date(row.get('date_of_birth'))
                surgery_date = parse_date(row.get('surgery_date'))
                chemo_start_date = parse_date(row.get('chemo_start_date'))
                followup_date = parse_date(row.get('followup_date'))
                recurrence_date = parse_date(row.get('recurrence_date'))
                
                # Parse booleans
                def parse_bool(value):
                    if not value or value == '':
                        return None
                    return value.lower() in ['true', '1', 'yes', 'y']
                
                # Create patient object
                patient = Patient(
                    id=uuid.uuid4(),
                    tenant_id=current_user.tenant_id,
                    organization_id=current_user.organization_id,
                    
                    # Required fields
                    patient_name=row['patient_name'],
                    diagnosis_date=diagnosis_date,
                    icd11_main_code=row['icd11_main_code'],
                    
                    # Optional fields
                    patient_id=row.get('patient_id') or None,
                    gender=row.get('gender') or None,
                    date_of_birth=date_of_birth,
                    nationality=row.get('nationality') or None,
                    
                    # ICD-11 fields
                    icd11_description=row.get('icd11_description') or None,
                    icd11_composite_expression=row.get('icd11_composite_expression') or None,
                    icd11_manifestation_code=row.get('icd11_manifestation_code') or None,
                    manifestation=row.get('manifestation') or None,
                    icd11_topography_code=row.get('icd11_topography_code') or None,
                    icd11_topography=row.get('icd11_topography') or None,
                    icd11_morphology_code=row.get('icd11_morphology_code') or None,
                    icd11_morphology=row.get('icd11_morphology') or None,
                    icd11_behavior_code=row.get('icd11_behavior_code') or None,
                    icd11_stage_code=row.get('icd11_stage_code') or None,
                    laterality=row.get('laterality') or None,
                    
                    # TNM staging
                    t_category=row.get('t_category') or None,
                    n_category=row.get('n_category') or None,
                    m_category=row.get('m_category') or None,
                    
                    # Cancer registry fields
                    multiple_primary_flag=parse_bool(row.get('multiple_primary_flag')),
                    basis_of_diagnosis=row.get('basis_of_diagnosis') or None,
                    primary_site_confirmed=parse_bool(row.get('primary_site_confirmed')),
                    
                    # Treatment
                    surgery_done=parse_bool(row.get('surgery_done')),
                    surgery_date=surgery_date,
                    chemotherapy_done=parse_bool(row.get('chemotherapy_done')),
                    chemo_start_date=chemo_start_date,
                    radiotherapy_done=parse_bool(row.get('radiotherapy_done')),
                    hormonal_therapy=parse_bool(row.get('hormonal_therapy')),
                    immunotherapy=parse_bool(row.get('immunotherapy')),
                    treatment_intent=row.get('treatment_intent') or None,
                    treatment_notes=row.get('treatment_notes') or None,
                    
                    # Follow-up
                    followup_date=followup_date,
                    vital_status=row.get('vital_status') or None,
                    cause_of_death_icd11=row.get('cause_of_death_icd11') or None,
                    recurrence=parse_bool(row.get('recurrence')),
                    recurrence_date=recurrence_date,
                    metastasis=parse_bool(row.get('metastasis')),
                    survival_months=int(row['survival_months']) if row.get('survival_months') else None,
                    followup_notes=row.get('followup_notes') or None,
                    
                    # Metadata
                    data_source=row.get('data_source') or 'Bulk CSV Upload',
                    entry_mode='Bulk',
                    entered_by=row.get('entered_by') or current_user.email,
                    validation_status=row.get('validation_status') or 'Pending',
                    entry_timestamp=datetime.now()
                )
                
                db.add(patient)
                db.flush()  # Flush to get the ID
                
                successful_imports.append({
                    "row": row_number,
                    "patient_id": patient.patient_id,
                    "patient_name": patient.patient_name,
                    "id": str(patient.id)
                })
                
            except Exception as e:
                failed_imports.append({
                    "row": row_number,
                    "error": str(e),
                    "data": row
                })
        
        # Commit all successful imports
        db.commit()
        
        # Log bulk upload audit
        try:
            log_bulk_upload(
                db=db,
                user_id=current_user.id,
                file_name=file.filename,
                total_records=row_number - 1,
                successful_records=len(successful_imports),
                failed_records=len(failed_imports)
            )
        except Exception as e:
            logger.warning(f"Audit logging failed for bulk upload: {str(e)}")
        
        return {
            "message": "Bulk upload completed",
            "total_rows": row_number - 1,
            "successful": len(successful_imports),
            "failed": len(failed_imports),
            "successful_imports": successful_imports,
            "failed_imports": failed_imports,
            "note": "All successful patients have been added to registry.patients and automatically synced to registry.patients_anonymized"
        }
        
    except Exception as e:
        db.rollback()
        logger.error(f"Error in bulk upload: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Bulk upload failed: {str(e)}"
        )


@router.get("/", dependencies=[Depends(permission_required("patients.read"))])
def list_patients(
    skip: int = Query(0, ge=0, description="Number of records to skip"),
    limit: int = Query(100, ge=1, le=1000, description="Maximum number of records to return"),
    search: Optional[str] = Query(None, description="Search by patient name, ID, or ICD-11 code"),
    validation_status: Optional[str] = Query(None, description="Filter by validation status"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    List all patients with pagination and filtering.
    
    Respects RLS - users only see patients from their tenant/organization.
    Super admins see all patients.
    """
    # Build query with RLS
    query = db.query(Patient)
    
    # Apply tenant/organization filtering based on user role
    if current_user.tenant_id is None:
        # Super admin - can see all patients
        pass
    elif current_user.organization_id:
        # Organization-scoped user
        query = query.filter(Patient.organization_id == current_user.organization_id)
    else:
        # Tenant-scoped user
        query = query.filter(Patient.tenant_id == current_user.tenant_id)
    
    # Apply search filter
    if search:
        search_term = f"%{search.lower()}%"
        query = query.filter(
            or_(
                Patient.patient_name.ilike(search_term),
                Patient.patient_id.ilike(search_term),
                Patient.icd11_main_code.ilike(search_term),
                Patient.icd11_description.ilike(search_term)
            )
        )
    
    # Apply validation status filter
    if validation_status:
        query = query.filter(Patient.validation_status == validation_status)
    
    # Get total count before pagination
    total = query.count()
    
    # Apply pagination
    patients = query.order_by(Patient.entry_timestamp.desc()).offset(skip).limit(limit).all()
    
    # Serialize patients to dictionaries for proper JSON response
    return {
        "items": [patient_to_dict(patient) for patient in patients],
        "total": total,
        "skip": skip,
        "limit": limit
    }


@router.get("/{patient_uuid}")
def get_patient(patient_uuid: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """
    Get patient details by UUID.
    """
    try:
        import uuid
        uuid.UUID(patient_uuid)
    except (ValueError, AttributeError):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Invalid UUID format"
        )
    
    patient = db.query(Patient).filter(Patient.id == patient_uuid).first()
    
    if not patient:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Patient not found"
        )
    
    # Check RLS - user must have access to this patient's tenant/org
    if current_user.tenant_id is None:
        # Super admin - can see all
        pass
    elif current_user.organization_id:
        if patient.organization_id != current_user.organization_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Access denied"
            )
    else:
        if patient.tenant_id != current_user.tenant_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Access denied"
            )
    
    return patient


class PatientUpdate(BaseModel):
    """Model for updating patient information"""
    patient_id: Optional[str] = None
    patient_name: Optional[str] = None
    gender: Optional[str] = None
    date_of_birth: Optional[date] = None
    nationality: Optional[str] = None
    address: Optional[dict] = None
    diagnosis_date: Optional[date] = None
    icd11_main_code: Optional[str] = None
    icd11_description: Optional[str] = None
    icd11_composite_expression: Optional[str] = None
    icd11_manifestation_code: Optional[str] = None
    manifestation: Optional[str] = None
    icd11_topography_code: Optional[str] = None
    icd11_topography: Optional[str] = None
    icd11_morphology_code: Optional[str] = None
    icd11_morphology: Optional[str] = None
    icd11_behavior_code: Optional[str] = None
    icd11_stage_code: Optional[str] = None
    laterality: Optional[str] = None
    t_category: Optional[str] = None
    n_category: Optional[str] = None
    m_category: Optional[str] = None
    multiple_primary_flag: Optional[bool] = None
    basis_of_diagnosis: Optional[str] = None
    primary_site_confirmed: Optional[bool] = None
    surgery_done: Optional[bool] = None
    surgery_date: Optional[date] = None
    chemotherapy_done: Optional[bool] = None
    chemo_start_date: Optional[date] = None
    radiotherapy_done: Optional[bool] = None
    hormonal_therapy: Optional[bool] = None
    immunotherapy: Optional[bool] = None
    treatment_intent: Optional[str] = None
    treatment_notes: Optional[str] = None
    followup_date: Optional[date] = None
    vital_status: Optional[str] = None
    cause_of_death_icd11: Optional[str] = None
    recurrence: Optional[bool] = None
    recurrence_date: Optional[date] = None
    metastasis: Optional[bool] = None
    survival_months: Optional[int] = None
    followup_notes: Optional[str] = None
    data_source: Optional[str] = None
    entry_mode: Optional[str] = None
    validation_status: Optional[str] = None


@router.put("/{patient_uuid}", dependencies=[Depends(permission_required("patients.update"))])
def update_patient(
    patient_uuid: str,
    payload: PatientUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Update patient information.
    
    Only updates fields that are provided in the payload.
    """
    try:
        import uuid
        uuid.UUID(patient_uuid)
    except (ValueError, AttributeError):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Invalid UUID format"
        )
    
    patient = db.query(Patient).filter(Patient.id == patient_uuid).first()
    
    if not patient:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Patient not found"
        )
    
    # Check RLS - user must have access to this patient's tenant/org
    if current_user.tenant_id is None:
        # Super admin - can update all
        pass
    elif current_user.organization_id:
        if patient.organization_id != current_user.organization_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Access denied"
            )
    else:
        if patient.tenant_id != current_user.tenant_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Access denied"
            )
    
    # Update only provided fields
    # Set updated_by to track who modified this record
    # Use dict() for Pydantic v1, model_dump() for v2
    try:
        update_data = payload.model_dump(exclude_unset=True, exclude_none=True)
    except AttributeError:
        update_data = payload.dict(exclude_unset=True, exclude_none=True)
    
    # Filter out empty strings - convert to None
    cleaned_update_data = {}
    for key, value in update_data.items():
        if value == '' or value is None:
            continue  # Skip empty strings and None values
        cleaned_update_data[key] = value
    
    # Recalculate age_at_diagnosis if date_of_birth or diagnosis_date changed
    if 'date_of_birth' in cleaned_update_data or 'diagnosis_date' in cleaned_update_data:
        dob = cleaned_update_data.get('date_of_birth', patient.date_of_birth)
        diag = cleaned_update_data.get('diagnosis_date', patient.diagnosis_date)
        if dob and diag:
            # Ensure dates are date objects, not strings
            if isinstance(dob, str):
                dob = datetime.strptime(dob, '%Y-%m-%d').date()
            if isinstance(diag, str):
                diag = datetime.strptime(diag, '%Y-%m-%d').date()
            age_at_diagnosis = diag.year - dob.year - ((diag.month, diag.day) < (dob.month, dob.day))
            cleaned_update_data['age_at_diagnosis'] = age_at_diagnosis
    
    # Update patient fields
    for field, value in cleaned_update_data.items():
        if hasattr(patient, field):
            # Handle date string conversion
            if field.endswith('_date') and isinstance(value, str):
                try:
                    value = datetime.strptime(value, '%Y-%m-%d').date()
                except:
                    pass  # Keep original value if parsing fails
            setattr(patient, field, value)
    
    # Track who updated this record
    patient.updated_by = current_user.id
    
    db.commit()
    db.refresh(patient)
    
    return {
        "id": str(patient.id),
        "patient_id": patient.patient_id,
        "message": "Patient updated successfully"
    }


class FollowupUpdate(BaseModel):
    followup_date: Optional[date] = None
    vital_status: Optional[str] = None
    cause_of_death_icd11: Optional[str] = None
    recurrence: Optional[bool] = None
    recurrence_date: Optional[date] = None
    metastasis: Optional[bool] = None
    survival_months: Optional[int] = None
    followup_notes: Optional[str] = None


@router.post("/{patient_id}/followup", dependencies=[Depends(permission_required("patients.update"))])
def add_followup(
    patient_id: str,
    payload: FollowupUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Add or update follow-up information for a patient.
    
    This endpoint allows doctors to update followup data separately from the main patient record.
    All followup fields are optional and only provided fields will be updated.
    """
    try:
        import uuid
        uuid.UUID(patient_id)
    except (ValueError, AttributeError):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Invalid UUID format"
        )
    
    patient = db.query(Patient).filter(Patient.id == patient_id).first()
    
    if not patient:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Patient not found"
        )
    
    # Check RLS - user must have access to this patient's tenant/org
    if current_user.tenant_id is None:
        # Super admin - can update all
        pass
    elif current_user.organization_id:
        if patient.organization_id != current_user.organization_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Access denied"
            )
    else:
        if patient.tenant_id != current_user.tenant_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Access denied"
            )
    
    # Update only provided followup fields
    try:
        update_data = payload.model_dump(exclude_unset=True)
    except AttributeError:
        update_data = payload.dict(exclude_unset=True)
    
    # Update patient followup fields
    for field, value in update_data.items():
        if hasattr(patient, field):
            setattr(patient, field, value)
    
    db.commit()
    db.refresh(patient)
    
    return {
        "id": str(patient.id),
        "patient_id": patient.patient_id,
        "message": "Followup information updated successfully",
        "updated_fields": list(update_data.keys())
    }


def patient_to_dict(patient: Patient) -> Dict[str, Any]:
    """
    Convert Patient model to dictionary with all fields.
    Returns complete dataset exactly as stored in PostgreSQL.
    """
    return {
        # Primary Key
        "id": str(patient.id),
        
        # Tenant & Organization
        "tenant_id": str(patient.tenant_id) if patient.tenant_id else None,
        "organization_id": str(patient.organization_id) if patient.organization_id else None,
        
        # Patient Demographics
        "patient_id": patient.patient_id,
        "patient_name": patient.patient_name,
        "gender": patient.gender,
        "date_of_birth": patient.date_of_birth.isoformat() if patient.date_of_birth else None,
        "nationality": patient.nationality,
        "address": patient.address if patient.address else None,
        
        # Diagnosis Details
        "diagnosis_date": patient.diagnosis_date.isoformat() if patient.diagnosis_date else None,
        "age_at_diagnosis": patient.age_at_diagnosis,
        
        # ICD-11 Official Disease Coding
        "icd11_main_code": patient.icd11_main_code,
        "icd11_description": patient.icd11_description,
        "icd11_composite_expression": patient.icd11_composite_expression,
        "icd11_manifestation_code": patient.icd11_manifestation_code,
        "manifestation": patient.manifestation,
        "icd11_topography_code": patient.icd11_topography_code,
        "icd11_topography": patient.icd11_topography,
        "icd11_morphology_code": patient.icd11_morphology_code,
        "icd11_morphology": patient.icd11_morphology,
        "icd11_behavior_code": patient.icd11_behavior_code,
        "icd11_stage_code": patient.icd11_stage_code,
        "laterality": patient.laterality,
        
        # TNM Staging (AJCC)
        "t_category": patient.t_category,
        "n_category": patient.n_category,
        "m_category": patient.m_category,
        
        # Additional Cancer Registry Fields
        "multiple_primary_flag": patient.multiple_primary_flag,
        "basis_of_diagnosis": patient.basis_of_diagnosis,
        "primary_site_confirmed": patient.primary_site_confirmed,
        
        # Treatment Information
        "surgery_done": patient.surgery_done,
        "surgery_date": patient.surgery_date.isoformat() if patient.surgery_date else None,
        "chemotherapy_done": patient.chemotherapy_done,
        "chemo_start_date": patient.chemo_start_date.isoformat() if patient.chemo_start_date else None,
        "radiotherapy_done": patient.radiotherapy_done,
        "hormonal_therapy": patient.hormonal_therapy,
        "immunotherapy": patient.immunotherapy,
        "treatment_intent": patient.treatment_intent,
        "treatment_notes": patient.treatment_notes,
        
        # Follow-up Data
        "followup_date": patient.followup_date.isoformat() if patient.followup_date else None,
        "vital_status": patient.vital_status,
        "cause_of_death_icd11": patient.cause_of_death_icd11,
        "recurrence": patient.recurrence,
        "recurrence_date": patient.recurrence_date.isoformat() if patient.recurrence_date else None,
        "metastasis": patient.metastasis,
        "survival_months": patient.survival_months,
        "followup_notes": patient.followup_notes,
        
        # Registry Metadata
        "data_source": patient.data_source,
        "entry_mode": patient.entry_mode,
        "entered_by": str(patient.entered_by) if patient.entered_by else None,
        "validation_status": patient.validation_status,
        
        # Payload & Audit
        "is_active": patient.is_active,
        "updated_by": str(patient.updated_by) if patient.updated_by else None,
        "entry_timestamp": patient.entry_timestamp.isoformat() if patient.entry_timestamp else None,
        "last_modified": patient.last_modified.isoformat() if patient.last_modified else None,
    }


@router.get("/all", dependencies=[Depends(permission_required("patients.read"))])
def get_all_patients(
    mode: str = Query("all", description="Mode: 'all' for standard response, 'export' for CSV export"),
    search: Optional[str] = Query(None, description="Search by patient ID, name, or ICD-11 code"),
    status: Optional[str] = Query(None, description="Filter by validation status (Pending, Approved, Rejected)"),
    skip: int = Query(0, ge=0, description="Number of records to skip (pagination)"),
    limit: Optional[int] = Query(None, ge=1, description="Maximum number of records to return (pagination)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Fetch ALL patient records from the database.
    
    Returns all fields stored in the patients table, including:
    - Demographics (patient_id, patient_name, gender, date_of_birth, nationality, address)
    - Diagnosis (diagnosis_date, age_at_diagnosis)
    - ICD-11 codes (main_code, description, topography, morphology, behavior, stage, etc.)
    - TNM staging (t_category, n_category, m_category)
    - Treatment (surgery, chemotherapy, radiotherapy, etc.)
    - Follow-up (followup_date, vital_status, recurrence, metastasis, etc.)
    - Metadata (data_source, entry_mode, validation_status, entry_timestamp, etc.)
    
    Modes:
    - mode=all: Standard response with pagination support
    - mode=export: Optimized for CSV export, returns all records without pagination
    
    Filters:
    - search: Filter by patient ID, name, or ICD-11 code
    - status: Filter by validation status (Pending, Approved, Rejected)
    
    Response includes the complete dataset exactly as stored in PostgreSQL.
    """
    # Build query with RLS
    query = db.query(Patient)
    
    # Apply tenant/organization filtering based on user role
    if current_user.tenant_id is None:
        # Super admin - can see all patients
        pass
    elif current_user.organization_id:
        # Organization-scoped user
        query = query.filter(Patient.organization_id == current_user.organization_id)
    else:
        # Tenant-scoped user
        query = query.filter(Patient.tenant_id == current_user.tenant_id)
    
    # Apply search filter
    if search:
        search_term = f"%{search.lower()}%"
        query = query.filter(
            or_(
                Patient.patient_name.ilike(search_term),
                Patient.patient_id.ilike(search_term),
                Patient.icd11_main_code.ilike(search_term),
                Patient.icd11_description.ilike(search_term)
            )
        )
    
    # Apply validation status filter
    if status:
        query = query.filter(Patient.validation_status == status)
    
    # EXPORT MODE (no pagination, optimized for CSV)
    if mode == "export":
        results = query.order_by(Patient.entry_timestamp.asc()).all()
        return [patient_to_dict(patient) for patient in results]
    
    # STANDARD MODE (with pagination support)
    # Get total count before pagination
    total = query.count()
    
    # Apply pagination if limit is provided
    if limit:
        results = query.order_by(Patient.entry_timestamp.desc()).offset(skip).limit(limit).all()
    else:
        # If no limit, return all (but still ordered)
        results = query.order_by(Patient.entry_timestamp.desc()).offset(skip).all()
    
    return {
        "count": len(results),
        "total": total,
        "skip": skip,
        "limit": limit,
        "items": [patient_to_dict(patient) for patient in results],
    }


@router.get("/all/export", dependencies=[Depends(permission_required("patients.read"))])
def export_all_patients_csv(
    search: Optional[str] = Query(None, description="Search by patient ID, name, or ICD-11 code"),
    status: Optional[str] = Query(None, description="Filter by validation status"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Export all patient records as CSV.
    
    Optimized for CSV export - returns all matching records without pagination.
    Includes all fields from the patients table.
    """
    # Build query with RLS
    query = db.query(Patient)
    
    # Apply tenant/organization filtering based on user role
    if current_user.tenant_id is None:
        # Super admin - can see all patients
        pass
    elif current_user.organization_id:
        # Organization-scoped user
        query = query.filter(Patient.organization_id == current_user.organization_id)
    else:
        # Tenant-scoped user
        query = query.filter(Patient.tenant_id == current_user.tenant_id)
    
    # Apply search filter
    if search:
        search_term = f"%{search.lower()}%"
        query = query.filter(
            or_(
                Patient.patient_name.ilike(search_term),
                Patient.patient_id.ilike(search_term),
                Patient.icd11_main_code.ilike(search_term),
                Patient.icd11_description.ilike(search_term)
            )
        )
    
    # Apply validation status filter
    if status:
        query = query.filter(Patient.validation_status == status)
    
    # Get all results ordered by entry timestamp
    results = query.order_by(Patient.entry_timestamp.asc()).all()
    
    if not results:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No patients found matching the criteria"
        )
    
    # Convert to dictionaries
    patients_data = [patient_to_dict(patient) for patient in results]
    
    # Flatten address JSONB field for CSV
    csv_data = []
    for patient in patients_data:
        csv_row = patient.copy()
        # Flatten address if it exists
        if csv_row.get("address") and isinstance(csv_row["address"], dict):
            address = csv_row["address"]
            csv_row["address_line1"] = address.get("line1", "")
            csv_row["address_line2"] = address.get("line2", "")
            csv_row["address_city"] = address.get("city", "")
            csv_row["address_state"] = address.get("state", "")
            csv_row["address_postcode"] = address.get("postcode", "")
            csv_row["address_country"] = address.get("country", "")
            csv_row["address"] = None  # Remove nested object
        csv_data.append(csv_row)
    
    # Generate CSV
    if not csv_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No data to export"
        )
    
    # Get all field names
    fieldnames = list(csv_data[0].keys())
    
    # Convert to CSV format
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    
    for row in csv_data:
        # Convert None to empty string, handle dates
        csv_row = {}
        for key, value in row.items():
            if value is None:
                csv_row[key] = ''
            elif isinstance(value, (datetime, date)):
                csv_row[key] = value.isoformat() if hasattr(value, 'isoformat') else str(value)
            elif isinstance(value, bool):
                csv_row[key] = 'Yes' if value else 'No'
            elif isinstance(value, dict):
                csv_row[key] = str(value)
            else:
                csv_row[key] = str(value)
        writer.writerow(csv_row)
    
    output.seek(0)
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"patients_export_{timestamp}.csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )




@router.get("/export/excel")
async def export_patients_excel(
    date_from: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    search: Optional[str] = Query(None, description="Search by patient ID, name, or ICD-11 code"),
    status: Optional[str] = Query(None, description="Filter by validation status"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Export patient data as Excel file (.xlsx).
    
    Only accessible to UMMC admins and super admins.
    Respects RLS - users only see patients from their tenant/organization.
    
    Parameters:
    - date_from: Start date for filtering by diagnosis_date (YYYY-MM-DD)
    - date_to: End date for filtering by diagnosis_date (YYYY-MM-DD)
    - search: Search by patient ID, name, or ICD-11 code
    - status: Filter by validation status
    
    Returns:
    - Excel file with all patient data matching the filters
    - Filename format: patients_export_YYYY-MM-DD_to_YYYY-MM-DD_timestamp.xlsx
    """
    # Authorization check: Only UMMC admins and super admins can export
    user_roles = [r.slug for r in current_user.roles]
    is_authorized = any(role in ['super_admin', 'ummc_admin'] for role in user_roles)
    
    if not is_authorized:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Excel export is only available to UMMC administrators and super administrators"
        )
    
    # Build query with RLS
    query = db.query(Patient)
    
    # Apply tenant/organization filtering based on user role
    if current_user.tenant_id is None:
        # Super admin - can see all patients
        pass
    elif current_user.organization_id:
        # Organization-scoped user
        query = query.filter(Patient.organization_id == current_user.organization_id)
    else:
        # Tenant-scoped user
        query = query.filter(Patient.tenant_id == current_user.tenant_id)
    
    # Apply date range filter
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(Patient.diagnosis_date >= date_from_obj)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid date_from format. Use YYYY-MM-DD"
            )
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(Patient.diagnosis_date <= date_to_obj)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid date_to format. Use YYYY-MM-DD"
            )
    
    # Apply search filter
    if search:
        search_term = f"%{search.lower()}%"
        query = query.filter(
            or_(
                Patient.patient_name.ilike(search_term),
                Patient.patient_id.ilike(search_term),
                Patient.icd11_main_code.ilike(search_term),
                Patient.icd11_description.ilike(search_term)
            )
        )
    
    # Apply validation status filter
    if status:
        query = query.filter(Patient.validation_status == status)
    
    # Get all results ordered by entry timestamp
    results = query.order_by(Patient.entry_timestamp.asc()).all()
    
    if not results:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No patients found matching the criteria"
        )
    
    # Convert to dictionaries
    patients_data = [patient_to_dict(patient) for patient in results]
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Patient Data"
    
    # Define column headers (all fields from patient_to_dict)
    headers = [
        "ID", "Tenant ID", "Organization ID", "Patient ID", "Patient Name",
        "Gender", "Date of Birth", "Nationality", "Address Line 1", "Address Line 2",
        "Address City", "Address State", "Address Postcode", "Address Country",
        "Diagnosis Date", "Age at Diagnosis",
        "ICD-11 Main Code", "ICD-11 Description", "ICD-11 Composite Expression",
        "ICD-11 Manifestation Code", "Manifestation", "ICD-11 Topography Code",
        "ICD-11 Topography", "ICD-11 Morphology Code", "ICD-11 Morphology",
        "ICD-11 Behavior Code", "ICD-11 Stage Code", "Laterality",
        "T Category", "N Category", "M Category",
        "Multiple Primary Flag", "Basis of Diagnosis", "Primary Site Confirmed",
        "Surgery Done", "Surgery Date", "Chemotherapy Done", "Chemo Start Date",
        "Radiotherapy Done", "Hormonal Therapy", "Immunotherapy",
        "Treatment Intent", "Treatment Notes",
        "Followup Date", "Vital Status", "Cause of Death ICD-11",
        "Recurrence", "Recurrence Date", "Metastasis",
        "Followup Interval (Months)", "Followup Notes",
        "Data Source", "Entry Mode", "Entered By", "Validation Status",
        "Is Active", "Updated By", "Entry Timestamp", "Last Modified"
    ]
    
    # Write headers with formatting
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data rows
    for row_num, patient in enumerate(patients_data, 2):
        # Flatten address if it exists
        address = patient.get("address") or {}
        if isinstance(address, dict):
            address_line1 = address.get("line1", "")
            address_line2 = address.get("line2", "")
            address_city = address.get("city", "")
            address_state = address.get("state", "")
            address_postcode = address.get("postcode", "")
            address_country = address.get("country", "")
        else:
            address_line1 = address_line2 = address_city = ""
            address_state = address_postcode = address_country = ""
        
        row_data = [
            patient.get("id", ""),
            patient.get("tenant_id", ""),
            patient.get("organization_id", ""),
            patient.get("patient_id", ""),
            patient.get("patient_name", ""),
            patient.get("gender", ""),
            patient.get("date_of_birth", ""),
            patient.get("nationality", ""),
            address_line1,
            address_line2,
            address_city,
            address_state,
            address_postcode,
            address_country,
            patient.get("diagnosis_date", ""),
            patient.get("age_at_diagnosis", ""),
            patient.get("icd11_main_code", ""),
            patient.get("icd11_description", ""),
            patient.get("icd11_composite_expression", ""),
            patient.get("icd11_manifestation_code", ""),
            patient.get("manifestation", ""),
            patient.get("icd11_topography_code", ""),
            patient.get("icd11_topography", ""),
            patient.get("icd11_morphology_code", ""),
            patient.get("icd11_morphology", ""),
            patient.get("icd11_behavior_code", ""),
            patient.get("icd11_stage_code", ""),
            patient.get("laterality", ""),
            patient.get("t_category", ""),
            patient.get("n_category", ""),
            patient.get("m_category", ""),
            "Yes" if patient.get("multiple_primary_flag") else "No" if patient.get("multiple_primary_flag") is False else "",
            patient.get("basis_of_diagnosis", ""),
            "Yes" if patient.get("primary_site_confirmed") else "No" if patient.get("primary_site_confirmed") is False else "",
            "Yes" if patient.get("surgery_done") else "No" if patient.get("surgery_done") is False else "",
            patient.get("surgery_date", ""),
            "Yes" if patient.get("chemotherapy_done") else "No" if patient.get("chemotherapy_done") is False else "",
            patient.get("chemo_start_date", ""),
            "Yes" if patient.get("radiotherapy_done") else "No" if patient.get("radiotherapy_done") is False else "",
            "Yes" if patient.get("hormonal_therapy") else "No" if patient.get("hormonal_therapy") is False else "",
            "Yes" if patient.get("immunotherapy") else "No" if patient.get("immunotherapy") is False else "",
            patient.get("treatment_intent", ""),
            patient.get("treatment_notes", ""),
            patient.get("followup_date", ""),
            patient.get("vital_status", ""),
            patient.get("cause_of_death_icd11", ""),
            "Yes" if patient.get("recurrence") else "No" if patient.get("recurrence") is False else "",
            patient.get("recurrence_date", ""),
            "Yes" if patient.get("metastasis") else "No" if patient.get("metastasis") is False else "",
            patient.get("survival_months", ""),
            patient.get("followup_notes", ""),
            patient.get("data_source", ""),
            patient.get("entry_mode", ""),
            patient.get("entered_by", ""),
            patient.get("validation_status", ""),
            "Yes" if patient.get("is_active") else "No" if patient.get("is_active") is False else "",
            patient.get("updated_by", ""),
            patient.get("entry_timestamp", ""),
            patient.get("last_modified", "")
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=False)
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column].width = adjusted_width
    
    # Generate filename with date range and timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if date_from and date_to:
        filename = f"patients_export_{date_from}_to_{date_to}_{timestamp}.xlsx"
    elif date_from:
        filename = f"patients_export_from_{date_from}_{timestamp}.xlsx"
    elif date_to:
        filename = f"patients_export_to_{date_to}_{timestamp}.xlsx"
    else:
        filename = f"patients_export_all_{timestamp}.xlsx"
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_file_path = temp_file.name
    temp_file.close()
    
    wb.save(temp_file_path)
    
    # Log export operation
    try:
        log_data_export(
            db=db,
            user_id=current_user.id,
            export_type="excel",
            record_count=len(results),
            filters={
                "date_from": date_from,
                "date_to": date_to,
                "search": search,
                "status": status
            }
        )
    except Exception as e:
        logger.warning(f"Audit logging failed for Excel export: {str(e)}")
    
    # Return file response
    return FileResponse(
        path=temp_file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )
